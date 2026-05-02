import argparse
from pathlib import Path
from typing import Optional

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


DEFAULT_LINKS_CSV = "outputs/sptulsian_ipo_analysis_blog_list_links.csv"
DEFAULT_OUTPUT_XLSX = "outputs/sptulsian_ipo_articles_text.xlsx"
USER_AGENT = "Mozilla/5.0"
OUTPUT_SHEET_NAME = "articles"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Grab article text/details from SPTulsian IPO analysis links."
    )
    parser.add_argument(
        "--links-file",
        default=DEFAULT_LINKS_CSV,
        help=f"CSV containing links. Default: {DEFAULT_LINKS_CSV}",
    )
    parser.add_argument(
        "--line-number",
        type=int,
        help="Fetch only one CSV row number (1-based). Good for quick debugging.",
    )
    parser.add_argument(
        "--start-line",
        type=int,
        help="Start CSV row number (1-based, inclusive) for ranged fetch.",
    )
    parser.add_argument(
        "--end-line",
        type=int,
        help="End CSV row number (1-based, inclusive) for ranged fetch.",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Fetch all rows from the links CSV.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        help="HTTP timeout in seconds. Default: 30",
    )
    parser.add_argument(
        "--output-file",
        default=DEFAULT_OUTPUT_XLSX,
        help=f"Output Excel file path. Default: {DEFAULT_OUTPUT_XLSX}",
    )
    return parser.parse_args()


def validate_selection_args(args: argparse.Namespace) -> None:
    if args.line_number is not None:
        if args.line_number <= 0:
            raise ValueError("--line-number must be >= 1.")
        if args.start_line is not None or args.end_line is not None or args.all:
            raise ValueError(
                "Use only one selection mode: --line-number OR --start-line/--end-line OR --all."
            )
        return

    if args.start_line is not None or args.end_line is not None:
        if args.all:
            raise ValueError(
                "Use only one selection mode: --line-number OR --start-line/--end-line OR --all."
            )
        if args.start_line is None or args.end_line is None:
            raise ValueError("Both --start-line and --end-line are required for range mode.")
        if args.start_line <= 0 or args.end_line <= 0:
            raise ValueError("--start-line and --end-line must be >= 1.")
        if args.start_line > args.end_line:
            raise ValueError("--start-line must be <= --end-line.")
        return

    if args.all:
        return

    raise ValueError(
        "Please choose one mode: --line-number, --start-line/--end-line, or --all."
    )


def get_article_text(main_div: Optional[BeautifulSoup]) -> str:
    if main_div is None:
        return ""

    content_div = main_div.select_one(
        "div.font-size-13.text-justify.width-100-percent.float-left"
    )
    if content_div is None:
        content_div = main_div

    text_parts = []
    seen = set()
    for p_tag in content_div.find_all("p"):
        p_text = p_tag.get_text(" ", strip=True)
        if p_text and p_text not in seen:
            seen.add(p_text)
            text_parts.append(p_text)

    return "\n".join(text_parts).strip()


def extract_article_details(url: str, timeout: int) -> dict:
    response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=timeout)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    main_div = soup.find("div", id="article-page-main-div")

    meta_title_tag = soup.find("meta", attrs={"name": "twitter:title"})
    meta_desc_tag = soup.find("meta", attrs={"name": "twitter:description"})

    meta_title = (meta_title_tag.get("content", "") if meta_title_tag else "").strip()
    meta_description = (meta_desc_tag.get("content", "") if meta_desc_tag else "").strip()

    h1_title = ""
    if main_div is not None:
        h1_tag = main_div.find("h1", class_="blog_article_title page-header-custom")
        if h1_tag is not None:
            h1_title = h1_tag.get_text(" ", strip=True)

    company_name = meta_title or h1_title
    article_text = get_article_text(main_div)

    return {
        "link": url,
        "company_name": company_name,
        "twitter_title": meta_title,
        "twitter_description": meta_description,
        "h1_title": h1_title,
        "article_text": article_text,
    }


def select_rows(df: pd.DataFrame, args: argparse.Namespace) -> pd.DataFrame:
    # Selection strategy:
    # - single line for fast validation and debugging,
    # - line range for controlled batches and resumability,
    # - all for full dataset processing.
    if args.line_number is not None:
        idx = args.line_number - 1
        if idx >= len(df):
            raise ValueError(f"--line-number {args.line_number} exceeds CSV length {len(df)}.")
        return df.iloc[[idx]].copy()

    if args.start_line is not None and args.end_line is not None:
        start_idx = args.start_line - 1
        end_idx = args.end_line - 1
        if start_idx >= len(df):
            raise ValueError(f"--start-line {args.start_line} exceeds CSV length {len(df)}.")
        end_idx = min(end_idx, len(df) - 1)
        return df.iloc[start_idx : end_idx + 1].copy()

    return df.copy()


def main() -> None:
    args = parse_args()
    validate_selection_args(args)

    df = pd.read_csv(args.links_file)
    if "href" not in df.columns:
        raise ValueError(f"'href' column not found in {args.links_file}.")
    if df.empty:
        print("No links found in CSV.")
        return

    selected = select_rows(df, args).reset_index(drop=False).rename(columns={"index": "source_index"})
    print(f"Selected {len(selected)} article link(s) from {len(df)} total rows.")
    results = []

    for _, row in selected.iterrows():
        source_line = int(row["source_index"]) + 1
        url = str(row["href"]).strip()
        if not url:
            continue

        print("\n" + "=" * 80)
        print(f"SOURCE_LINE: {source_line} for URL: {url}")
        #print(f"URL: {url}")
        try:
            details = extract_article_details(url, timeout=args.timeout)
        except Exception as exc:  # noqa: BLE001 - continue processing remaining links.
            print(f"ERROR: Failed to process article: {exc}")
            continue

        #print(f"COMPANY_NAME: {details['company_name']}")
        #print(f"TWITTER_TITLE: {details['twitter_title']}")
        #print(f"TWITTER_DESCRIPTION: {details['twitter_description']}")
        #print(f"H1_TITLE: {details['h1_title']}")
        #print("ARTICLE_TEXT:")
        #print(details["article_text"] if details["article_text"] else "(empty)")
        #print("=" * 80)
        results.append(
            {
                "company_name": details["company_name"],
                "link": details["link"],
                "mini_desc": details["twitter_description"],
                "article_text": details["article_text"],
            }
        )

    out_df = pd.DataFrame(results, columns=["company_name", "link", "mini_desc", "article_text"])
    output_path = Path(args.output_file)
    if output_path.parent:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    if out_df.empty:
        print("\nNo successful records found; nothing was written to Excel.")
        return

    if output_path.exists():
        workbook = load_workbook(output_path)
        startrow = 0
        if OUTPUT_SHEET_NAME in workbook.sheetnames:
            worksheet = workbook[OUTPUT_SHEET_NAME]
            startrow = worksheet.max_row
        workbook.close()

        with pd.ExcelWriter(
            output_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="overlay",
        ) as writer:
            out_df.to_excel(
                writer,
                sheet_name=OUTPUT_SHEET_NAME,
                index=False,
                header=(startrow == 0),
                startrow=startrow,
            )
    else:
        out_df.to_excel(output_path, sheet_name=OUTPUT_SHEET_NAME, index=False)

    print(
        f"\nSaved {len(out_df)} record(s) to Excel (append mode when file exists): {output_path}"
    )


if __name__ == "__main__":
    main()
